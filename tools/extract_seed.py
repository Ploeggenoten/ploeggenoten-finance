#!/usr/bin/env python3
"""Extract Ploeggenoten Excel data -> supabase/seed.sql"""
import openpyxl, datetime, json, re

FD = '/Users/tjeerdvanelk/Desktop/Facturatie_Dashboard_Ploeggenoten_bijgewerkt.xlsx'
OUT = '/Users/tjeerdvanelk/ploeggenoten-finance/supabase/seed.sql'

wb = openpyxl.load_workbook(FD, data_only=True)

def d(v):
    if isinstance(v, datetime.datetime): return v.date().isoformat()
    if isinstance(v, datetime.date): return v.isoformat()
    return None

def num(v):
    if v is None or v == '' or v == '—': return None
    try: return round(float(v), 2)
    except: return None

def esc(s):
    return "'" + str(s).replace("'", "''") + "'"

def sqlv(v):
    if v is None: return 'null'
    if isinstance(v, bool): return 'true' if v else 'false'
    if isinstance(v, (int, float)): return str(v)
    return esc(v)

# ---------- Plaatsingen ----------
ws = wb['Plaatsingen']
placements = {}
for row in ws.iter_rows(min_row=2, values_only=False):
    pid = row[0].value
    klant = row[1].value
    if not pid or not klant: continue
    placements[pid] = dict(
        id=pid, klant=str(klant).strip(), kandidaat=(row[2].value or '').strip(),
        functie=(row[3].value or '').strip(), fee_excl=num(row[4].value),
        contract_datum=d(row[5].value), eerste_factuurdatum=d(row[6].value),
        aantal_termijnen=int(row[7].value or 1), maanden_tussen=int(row[8].value or 1),
        betaaltermijn_dgn=int(row[9].value or 14),
        garantie_mnd=int(row[20].value or 0), gestopt_op=d(row[21].value),
        garantie_note=(str(row[22].value).strip() if row[22].value else None),
    )

# ---------- Facturatie: geplande datum + bedrag per (pid, termijn) ----------
ws = wb['Facturatie']
plan = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    pid, klant, kand, tnr = row[0], row[1], row[2], row[3]
    if not pid or tnr is None: continue
    plan[(pid, int(tnr))] = dict(
        bedrag_excl=num(row[4]), geplande_datum=d(row[6]),
        eff_status=(row[13] or '').strip() if row[13] else '',
        signaal=(row[14] or '').strip() if isinstance(row[14], str) else '',
    )

# ---------- Betaalstatus: werkelijke status per (pid, termijn) ----------
ws = wb['Betaalstatus']
status = {}
for row in ws.iter_rows(min_row=3, values_only=True):
    pid, tnr = row[0], row[1]
    if not pid or tnr is None: continue
    gef = str(row[5] or '').strip().lower() == 'ja'
    bet = str(row[7] or '').strip().lower() == 'ja'
    status[(pid, int(tnr))] = dict(
        gefactureerd=gef, factuurdatum=d(row[6]), betaald=bet, betaaldatum=d(row[8]),
        _status=(row[9] or '').strip() if row[9] else '',
    )

# ---------- Build installments ----------
installments = []
for (pid, tnr), st in sorted(status.items()):
    if pid not in placements: continue
    p = placements[pid]
    pl = plan.get((pid, tnr), {})
    if st['_status'].lower() == 'vervallen':
        s = 'vervallen'
    elif st['gefactureerd'] and st['betaald']:
        s = 'betaald'
    elif st['gefactureerd']:
        s = 'gefactureerd'
    else:
        s = 'te_factureren'
    bedrag = pl.get('bedrag_excl')
    if bedrag is None:
        bedrag = round((p['fee_excl'] or 0) / p['aantal_termijnen'], 2)
    installments.append(dict(
        placement_id=pid, termijn_nr=tnr, bedrag_excl=bedrag,
        geplande_datum=pl.get('geplande_datum'), status=s,
        factuurdatum=st['factuurdatum'], betaaldatum=st['betaaldatum'],
    ))

# sanity checks vs dashboard numbers
tot_pijplijn = sum(p['fee_excl'] or 0 for p in placements.values())
tot_gefact = sum(i['bedrag_excl'] for i in installments if i['status'] in ('gefactureerd','betaald'))
tot_open = sum(i['bedrag_excl'] for i in installments if i['status'] == 'gefactureerd')
tot_nog = sum(i['bedrag_excl'] for i in installments if i['status'] == 'te_factureren')
tot_verv = sum(i['bedrag_excl'] for i in installments if i['status'] == 'vervallen')
print(f'placements={len(placements)} installments={len(installments)}')
print(f'pijplijn={tot_pijplijn:.2f} (verwacht ~210847)')
print(f'gefactureerd={tot_gefact:.2f} open_niet_betaald={tot_open:.2f} (verwacht 17917)')
print(f'nog_te_factureren={tot_nog:.2f} (verwacht 46666) vervallen={tot_verv:.2f} (verwacht 16796)')

# ---------- Emit SQL ----------
L = []
L.append('-- Seed: historische data uit Excel (gegenereerd ' + datetime.date.today().isoformat() + ')')
L.append('-- Draai NA schema.sql. Idempotent: bestaande rijen worden overschreven.')
L.append('')
for p in sorted(placements.values(), key=lambda x: x['id']):
    cols = ['id','klant','kandidaat','functie','fee_excl','contract_datum','eerste_factuurdatum',
            'aantal_termijnen','maanden_tussen','betaaltermijn_dgn','garantie_mnd','gestopt_op','garantie_note','bron']
    vals = [sqlv(p['id']), sqlv(p['klant']), sqlv(p['kandidaat']), sqlv(p['functie']), sqlv(p['fee_excl']),
            sqlv(p['contract_datum']), sqlv(p['eerste_factuurdatum']), sqlv(p['aantal_termijnen']),
            sqlv(p['maanden_tussen']), sqlv(p['betaaltermijn_dgn']), sqlv(p['garantie_mnd']),
            sqlv(p['gestopt_op']), sqlv(p['garantie_note']), sqlv('excel')]
    L.append(f"insert into fin_placements ({','.join(cols)}) values ({','.join(vals)})")
    L.append("  on conflict (id) do update set " + ', '.join(f'{c}=excluded.{c}' for c in cols[1:]) + ';')
L.append('')
for i in installments:
    cols = ['placement_id','termijn_nr','bedrag_excl','geplande_datum','status','factuurdatum','betaaldatum']
    vals = [sqlv(i[c]) for c in cols]
    L.append(f"insert into fin_installments ({','.join(cols)}) values ({','.join(vals)})")
    L.append("  on conflict (placement_id, termijn_nr) do update set " + ', '.join(f'{c}=excluded.{c}' for c in cols[2:]) + ';')

with open(OUT, 'w') as f:
    f.write('\n'.join(L) + '\n')
print('wrote', OUT, len(L), 'lines')
